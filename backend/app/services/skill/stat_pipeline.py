"""统计分析 Skill：意图路由 + 白名单脚本执行。"""
from __future__ import annotations

import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Any

from app.logging_setup import get_logger
from app.core.paths import generated_root, skills_root

logger = get_logger("stat")

SCRIPTS_DIR = skills_root() / "statistical-analyst" / "scripts"
WHITELIST = {
    "hypothesis": "hypothesis_tester.py",
    "sample_size": "sample_size_calculator.py",
    "confidence_interval": "confidence_interval.py",
}
OUTPUT_CAP = 200_000
TIMEOUT_SEC = 30


def resolve_scripts_dir(scripts_dir: Path | None = None) -> Path:
    if scripts_dir and Path(scripts_dir).is_dir():
        return Path(scripts_dir)
    return SCRIPTS_DIR


def is_statistical_skill(agent) -> bool:
    wf = getattr(agent, "workflow", None) or {}
    if isinstance(wf, dict) and wf.get("kind") == "statistical-analyst":
        return True
    if (getattr(agent, "specialty", None) or "") == "统计分析":
        return True
    name = (getattr(agent, "name", None) or "").lower()
    return "statistical" in name or "统计分析" in (getattr(agent, "name", None) or "")


def prefer_current_user_text(text: str) -> str:
    """多轮会话会把历史拼进 input；参数抽取只看「当前用户」段，避免旧数字污染。"""
    t = text or ""
    marker = "## 当前用户"
    if marker in t:
        return t.split(marker, 1)[-1].strip()
    return t.strip()


def route_intent(text: str) -> str:
    t = (text or "").lower()
    if any(k in t for k in ("样本量", "sample size", "mde", "power", "功效", "要跑多久")):
        return "sample_size"
    if any(k in t for k in ("置信区间", "confidence interval", "margin of error", "ci for")):
        return "confidence_interval"
    if any(
        k in t
        for k in (
            "a/b",
            "假设检验",
            "z-test",
            "ztest",
            "t-test",
            "ttest",
            "chi",
            "转化",
            "对照",
            "实验组",
            "p-value",
            "p值",
            "excel",
            "xlsx",
            "汇总",
        )
    ):
        return "hypothesis"
    return "hypothesis"


def _ztest_args(
    cn: int, cx: int, tn: int, tx: int, *, alpha: float = 0.05
) -> list[str]:
    return [
        "--test",
        "ztest",
        "--control-n",
        str(cn),
        "--control-x",
        str(cx),
        "--treatment-n",
        str(tn),
        "--treatment-x",
        str(tx),
        "--alpha",
        str(alpha),
        "--format",
        "json",
    ]


def extract_alpha(text: str, default: float = 0.05) -> float:
    m = re.search(r"(?:α|alpha|显著性水平)\s*[=:：]?\s*(0?\.\d+)", text or "", re.I)
    if not m:
        return default
    try:
        v = float(m.group(1))
        return v if 0 < v < 1 else default
    except ValueError:
        return default


def extract_ab_counts_from_text(text: str) -> tuple[int, int, int, int] | None:
    """从自然语言抽出对照/实验 n、x。"""
    t = text or ""
    cn = re.search(r"(?:对照|控制|control)[^\d]{0,16}n\s*[=:：]?\s*(\d+)", t, re.I)
    cx = re.search(
        r"(?:对照|控制|control)[^\d]{0,24}(?:转化|成功|x|转化数)\s*[=:：]?\s*(\d+)", t, re.I
    )
    tn = re.search(r"(?:实验|处理|treatment)[^\d]{0,16}n\s*[=:：]?\s*(\d+)", t, re.I)
    tx = re.search(
        r"(?:实验|处理|treatment)[^\d]{0,24}(?:转化|成功|x|转化数)\s*[=:：]?\s*(\d+)",
        t,
        re.I,
    )
    if cn and cx and tn and tx:
        return int(cn.group(1)), int(cx.group(1)), int(tn.group(1)), int(tx.group(1))

    pairs = re.findall(
        r"n\s*=\s*(\d+)[^\d]{0,24}(?:转化|转化数|x)\s*[=:：]?\s*(\d+)", t, re.I
    )
    if len(pairs) >= 2:
        return int(pairs[0][0]), int(pairs[0][1]), int(pairs[1][0]), int(pairs[1][1])

    m = re.search(
        r"对照[^\d]*(\d+)[^\d]+(\d+).*?(?:实验|处理)[^\d]*(\d+)[^\d]+(\d+)",
        t,
        re.S,
    )
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))

    # 表格预览：对照 control 5000 250
    m2 = re.search(
        r"对照\s+\S+\s+(\d+)\s+(\d+).*?(?:实验|处理)\s+\S+\s+(\d+)\s+(\d+)",
        t,
        re.S | re.I,
    )
    if m2:
        return int(m2.group(1)), int(m2.group(2)), int(m2.group(3)), int(m2.group(4))
    return None


def extract_ab_counts_from_excel(path: Path) -> tuple[int, int, int, int] | None:
    """读取 Excel「AB汇总」或首个含对照/实验行的 Sheet。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("excel open fail path=%s err=%s", path, exc)
        return None

    try:
        sheets = list(wb.worksheets)
        preferred = [s for s in sheets if "汇总" in (s.title or "")] + sheets
        control: tuple[int, int] | None = None
        treatment: tuple[int, int] | None = None
        for sheet in preferred:
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if c is None else str(c).strip() for c in row]
                if not any(cells):
                    continue
                line = " ".join(cells).lower()
                # 找 n、x：优先第 3、4 列（我们生成的模板）
                nums = [int(float(x)) for x in cells if re.fullmatch(r"\d+(\.0+)?", x)]
                if ("对照" in cells[0] or "control" in line) and len(nums) >= 2:
                    control = (nums[0], nums[1])
                if (
                    "实验" in (cells[0] if cells else "")
                    or "treatment" in line
                    or "处理" in (cells[0] if cells else "")
                ) and len(nums) >= 2:
                    treatment = (nums[0], nums[1])
            if control and treatment:
                return control[0], control[1], treatment[0], treatment[1]
        return None
    finally:
        wb.close()


def extract_ab_from_uploads(files: list[dict] | None) -> tuple[int, int, int, int] | None:
    for f in files or []:
        path = Path(str(f.get("path") or ""))
        if path.suffix.lower() in {".xlsx", ".xls"} and path.is_file():
            got = extract_ab_counts_from_excel(path)
            if got:
                logger.info("ab from excel %s -> %s", path.name, got)
                return got
        preview = str(f.get("preview") or "")
        if preview:
            got = extract_ab_counts_from_text(preview)
            if got:
                logger.info("ab from preview %s -> %s", f.get("name"), got)
                return got
        # csv 路径
        if path.suffix.lower() == ".csv" and path.is_file():
            try:
                got = extract_ab_counts_from_text(path.read_text(encoding="utf-8-sig"))
                if got:
                    return got
            except Exception:  # noqa: BLE001
                pass
    return None


def build_cli_args(intent: str, text: str) -> list[str] | None:
    """从自然语言尽力抽出 CLI；失败返回 None。"""
    t = prefer_current_user_text(text)
    alpha = extract_alpha(t)
    if intent == "hypothesis":
        ab = extract_ab_counts_from_text(t)
        if ab:
            return _ztest_args(*ab, alpha=alpha)
        return None

    if intent == "sample_size":
        base = re.search(
            r"(?:baseline|基线|对照率|转化率)\s*[=:：约]?\s*(0?\.\d+|\d+(?:\.\d+)?%?)",
            t,
            re.I,
        )
        mde = re.search(
            r"(?:mde|最小可检测|相对(?:提升|mde)?)\s*[=:：]?\s*(0?\.\d+|\d+(?:\.\d+)?%?)",
            t,
            re.I,
        )
        # 相对提升 20%
        if not mde:
            mde = re.search(r"相对提升\s*(\d+(?:\.\d+)?)\s*%", t, re.I)
        if not base:
            base = re.search(r"基线转化率约?\s*(\d+(?:\.\d+)?)\s*%", t, re.I)
        power_m = re.search(r"(?:power|功效)\s*[=:：]?\s*(0?\.\d+)", t, re.I)
        power = power_m.group(1) if power_m else "0.8"
        if base and mde:
            def pct(s: str) -> str:
                s = s.strip().rstrip("%％")
                v = float(s)
                return str(v if v <= 1 else v / 100.0)

            return [
                "--test",
                "proportion",
                "--baseline",
                pct(base.group(1)),
                "--mde",
                pct(mde.group(1)),
                "--alpha",
                str(alpha),
                "--power",
                power,
                "--format",
                "json",
            ]
        return None

    if intent == "confidence_interval":
        # 实验组 … n=5000，转化 310
        n = re.search(
            r"(?:实验组|treatment)?[^\n]{0,20}\bn\s*[=:：]?\s*(\d+)", t, re.I
        )
        x = re.search(
            r"(?:转化|成功|x)\s*[=:：]?\s*(\d+)", t, re.I
        )
        conf_m = re.search(r"(?:置信度|confidence)\s*[=:：]?\s*(0?\.\d+)", t, re.I)
        # 95% 置信区间
        if not conf_m:
            conf_m = re.search(r"(\d+(?:\.\d+)?)\s*%\s*置信", t)
        if conf_m:
            raw = float(conf_m.group(1))
            confidence = str(raw if raw <= 1 else raw / 100.0)
        else:
            confidence = "0.95"
        if n and x:
            return [
                "--type",
                "proportion",
                "--n",
                n.group(1),
                "--x",
                x.group(1),
                "--confidence",
                confidence,
                "--format",
                "json",
            ]
        return None
    return None


def run_whitelisted_script(
    intent: str,
    cli_args: list[str],
    *,
    scripts_dir: Path | None = None,
) -> dict[str, Any]:
    script_name = WHITELIST.get(intent)
    if not script_name:
        raise ValueError(f"未知统计意图: {intent}")
    base = resolve_scripts_dir(scripts_dir)
    script = base / script_name
    if not script.is_file():
        raise FileNotFoundError(f"脚本不存在: {script}")
    cmd = [sys.executable, str(script), *cli_args]
    logger.info("run script=%s args=%s", script_name, " ".join(cli_args))
    proc = subprocess.run(
        cmd,
        cwd=str(base),
        capture_output=True,
        text=True,
        timeout=TIMEOUT_SEC,
        shell=False,
        check=False,
    )
    stdout = (proc.stdout or "")[:OUTPUT_CAP]
    stderr = (proc.stderr or "")[:OUTPUT_CAP]
    return {
        "script": script_name,
        "intent": intent,
        "cmd": " ".join(cmd[1:]),
        "exitCode": proc.returncode,
        "stdout": stdout,
        "stderr": stderr,
    }


def run_statistical_tools(
    user_text: str,
    *,
    uploaded_files: list[dict] | None = None,
    scripts_dir: Path | None = None,
) -> dict[str, Any]:
    source = prefer_current_user_text(user_text)
    intent = route_intent(source)
    args = build_cli_args(intent, source)

    # 假设检验：文本抽不到时，从 Excel/CSV 读 AB汇总
    if intent == "hypothesis" and not args:
        ab = extract_ab_from_uploads(uploaded_files)
        if ab:
            args = _ztest_args(*ab)

    if not args:
        return {
            "script": None,
            "intent": intent,
            "exitCode": None,
            "stdout": "",
            "stderr": "",
            "note": (
                "未能自动抽出脚本参数。请在提问中写明："
                "对照 n=… 转化 …；实验 n=… 转化 …；"
                "或上传含「AB汇总」的 Excel（对照/实验两行含样本量与转化数）。"
            ),
        }
    try:
        result = run_whitelisted_script(intent, args, scripts_dir=scripts_dir)
        if result.get("exitCode") not in (0, None):
            result["note"] = "脚本非零退出；请检查参数或查看 stderr"
        return result
    except Exception as exc:  # noqa: BLE001
        logger.error("stat script fail: %s", exc)
        return {
            "script": WHITELIST.get(intent),
            "intent": intent,
            "exitCode": 1,
            "stdout": "",
            "stderr": str(exc),
            "note": f"脚本执行异常: {exc}",
        }


def format_tool_result(trace: dict[str, Any]) -> str:
    if not trace:
        return "（无脚本结果）"
    parts = [
        f"intent: {trace.get('intent')}",
        f"script: {trace.get('script')}",
        f"exitCode: {trace.get('exitCode')}",
    ]
    if trace.get("note"):
        parts.append(f"note: {trace['note']}")
    if trace.get("stdout"):
        parts.append("stdout:\n" + trace["stdout"])
        # 提示模型：JSON 已给出正式推断结果
        if (trace.get("stdout") or "").lstrip().startswith("{"):
            parts.append(
                "重要：以上 stdout JSON 为权威统计结果；"
                "解读时必须使用其中的 p_value / ci95_diff / significant / effect_size_h，禁止写成 N/A 或声称脚本失败。"
            )
    if trace.get("stderr"):
        parts.append("stderr:\n" + trace["stderr"])
    if trace.get("intent") == "intro":
        parts.append("重要：intent=intro。介绍能力即可，禁止套用 Bottom Line / What 任务模板。")
    return "\n".join(parts)
