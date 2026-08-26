"""Excel 表格助手：DuckDB 分析、openpyxl 写出、matplotlib 出图。"""
from __future__ import annotations

import copy
import re
import uuid
from pathlib import Path
from typing import Any

from app.logging_setup import get_logger

logger = get_logger("excel")

ROOT = Path(__file__).resolve().parents[4]
GENERATED_DIR = ROOT / "static" / "generated"
TABULAR_SUFFIX = {".xlsx", ".xls", ".csv", ".tsv"}
_SQL_FORBIDDEN = re.compile(
    r"\b(insert|update|delete|drop|alter|attach|copy|pragma|install|export|create\s+or)\b",
    re.I,
)
_CJK_FONT = Path(r"C:\Windows\Fonts\msyh.ttc")


def _public_url(name: str) -> str:
    return f"/static/generated/{name}"


def sanitize_table_name(name: str) -> str:
    sanitized = re.sub(r"[^\w]", "_", name or "table", flags=re.UNICODE)
    if sanitized and sanitized[0].isdigit():
        sanitized = f"t_{sanitized}"
    return sanitized or "table"


def unique_table(name: str, used: set[str]) -> str:
    base = sanitize_table_name(name)
    out = base
    i = 1
    while out in used:
        out = f"{base}_{i}"
        i += 1
    used.add(out)
    return out


def is_numeric_type(dtype: str) -> bool:
    t = (dtype or "").upper()
    return any(
        k in t
        for k in (
            "INT",
            "DOUBLE",
            "FLOAT",
            "DECIMAL",
            "HUGE",
            "REAL",
            "NUM",
            "BIGINT",
        )
    )


def load_tables(files: list[dict]) -> tuple[Any, dict[str, str]]:
    import duckdb
    import pandas as pd

    con = duckdb.connect()
    table_map: dict[str, str] = {}
    used: set[str] = set()
    for src in files:
        path = Path(str(src.get("path") or ""))
        if not path.is_file():
            continue
        ext = path.suffix.lower()
        if ext in {".xlsx", ".xls"}:
            try:
                engine = "openpyxl" if ext == ".xlsx" else None
                sheets = pd.read_excel(path, sheet_name=None, engine=engine)
            except Exception as exc:  # noqa: BLE001
                logger.warning("read_excel fail %s %s", path, exc)
                continue
            for sheet_name, df in (sheets or {}).items():
                if df is None or df.empty:
                    continue
                df = coerce_numeric_columns(df.copy())
                df.columns = [str(c).strip() or f"col_{i}" for i, c in enumerate(df.columns)]
                df = attach_src_order(df)
                label = str(sheet_name)
                tbl = unique_table(label, used)
                con.register(f"_df_{tbl}", df)
                con.execute(f'CREATE TABLE "{tbl}" AS SELECT * FROM "_df_{tbl}"')
                table_map[label] = tbl
        elif ext in {".csv", ".tsv"}:
            sep = "\t" if ext == ".tsv" else ","
            try:
                df = pd.read_csv(path, sep=sep)
            except Exception as exc:  # noqa: BLE001
                logger.warning("read_csv fail %s %s", path, exc)
                continue
            if df is None or df.empty:
                continue
            df = coerce_numeric_columns(df)
            df.columns = [str(c).strip() or f"col_{i}" for i, c in enumerate(df.columns)]
            df = attach_src_order(df)
            label = path.stem
            tbl = unique_table(label, used)
            con.register(f"_df_{tbl}", df)
            con.execute(f'CREATE TABLE "{tbl}" AS SELECT * FROM "_df_{tbl}"')
            table_map[label] = tbl
    return con, table_map


def describe_tables(con, table_map: dict[str, str]) -> str:
    parts: list[str] = []
    for original, tbl in table_map.items():
        n = con.execute(f'SELECT COUNT(*) FROM "{tbl}"').fetchone()[0]
        cols = con.execute(f'DESCRIBE "{tbl}"').fetchall()
        parts.append(f'## 表 {original}（SQL: "{tbl}"，{n} 行）')
        for col in cols:
            if col[0] == "_src_order":
                continue
            parts.append(f"- {col[0]}: {col[1]}")
        vis = [c[0] for c in cols if c[0] != "_src_order"]
        sel = ", ".join(f'"{c}"' for c in vis) or "*"
        sample = con.execute(f'SELECT {sel} FROM "{tbl}" LIMIT 5').fetchdf()
        parts.append("样例（前 5 行）:")
        parts.append(sample.to_string(index=False))
        parts.append("")
    return "\n".join(parts).strip() or "未载入任何表格。"


def summarize_table(con, table_map: dict[str, str], table_hint: str = "") -> str:
    if not table_map:
        return "没有可汇总的表。"
    resolved = None
    if table_hint:
        for orig, tbl in table_map.items():
            if table_hint.lower() in orig.lower() or table_hint.lower() == tbl.lower():
                resolved = tbl
                break
    if not resolved:
        resolved = next(iter(table_map.values()))
    orig = next((k for k, v in table_map.items() if v == resolved), resolved)
    cols = con.execute(f'DESCRIBE "{resolved}"').fetchall()
    n = con.execute(f'SELECT COUNT(*) FROM "{resolved}"').fetchone()[0]
    parts = [f"## 描述统计：{orig}（{n} 行）"]
    for col_name, col_type, *_rest in cols:
        if col_name == "_src_order":
            continue
        if is_numeric_type(str(col_type)):
            row = con.execute(
                f"""
                SELECT COUNT("{col_name}"), AVG("{col_name}"), STDDEV_SAMP("{col_name}"),
                       MIN("{col_name}"), MAX("{col_name}"),
                       COUNT(*) - COUNT("{col_name}")
                FROM "{resolved}"
                """
            ).fetchone()
            parts.append(
                f"- {col_name}（数值）count={row[0]} mean={row[1]} std={row[2]} "
                f"min={row[3]} max={row[4]} nulls={row[5]}"
            )
        else:
            row = con.execute(
                f"""
                SELECT COUNT("{col_name}"), COUNT(DISTINCT "{col_name}"),
                       COUNT(*) - COUNT("{col_name}")
                FROM "{resolved}"
                """
            ).fetchone()
            parts.append(
                f"- {col_name}（类别）count={row[0]} unique={row[1]} nulls={row[2]}"
            )
    return "\n".join(parts)


def extract_sql(text: str) -> str | None:
    t = text or ""
    m = re.search(r"```sql\s*([\s\S]+?)```", t, re.I)
    blob = m.group(1).strip() if m else ""
    if not blob:
        m2 = re.search(r"((?:with|select)\s[\s\S]{8,4000})", t, re.I)
        if not m2:
            return None
        blob = m2.group(1).strip()
        blob = re.split(r"\n(?:请|然后|另外|## )", blob, maxsplit=1)[0].strip()
    blob = blob.rstrip(";").strip()
    if _SQL_FORBIDDEN.search(blob):
        return None
    if not re.search(r"\bselect\b", blob, re.I):
        return None
    return blob


def rewrite_sql(sql: str, table_map: dict[str, str]) -> str:
    out = sql
    for original, tbl in sorted(table_map.items(), key=lambda x: len(x[0]), reverse=True):
        if original != tbl:
            out = re.sub(rf"\b{re.escape(original)}\b", f'"{tbl}"', out)
    return out


def run_sql(con, sql: str, table_map: dict[str, str]) -> tuple[list[str], list[tuple], str]:
    modified = rewrite_sql(sql, table_map)
    try:
        result = con.execute(modified)
        columns = [d[0] for d in result.description]
        rows = result.fetchall()
        return columns, rows, ""
    except Exception as exc:  # noqa: BLE001
        avail = ", ".join(f'{o}→"{t}"' for o, t in table_map.items())
        return [], [], f"SQL 失败: {exc}\n可用表: {avail}"


def current_user_text(text: str) -> str:
    """只看本轮用户话；去掉会话里拼上的「先前对话」。"""
    t = text or ""
    marker = "## 当前用户"
    if marker in t:
        t = t.rsplit(marker, 1)[-1]
    t = re.split(r"\n## (?:当前上传|上传文件|先前对话)", t, maxsplit=1)[0]
    return t.strip()


def strip_table_from_prompt(text: str) -> str:
    """去掉已识别的表格行，只留指令，避免表头列名被当成用户指定的汇总字段。"""
    parsed = parse_pasted_table(text)
    lines: list[str] = []
    for ln in (text or "").splitlines():
        if _INSTRUCTION_RE.search(ln) and len(_split_row(ln)) <= 3:
            lines.append(ln)
            continue
        cells = _split_row(ln)
        if parsed and len(cells) == len(parsed[0]) and len(cells) >= 2:
            continue
        if cells and not _is_md_sep(cells):
            lines.append(ln)
    return "\n".join(lines).strip() or (text or "")


def match_column(text: str, columns: list[str]) -> str | None:
    t = strip_table_from_prompt(text)
    hits = [c for c in columns if c and str(c) in t]
    if hits:
        return max(hits, key=len)
    return None


_MONTH_EXTRACT = re.compile(r"(\d+)\s*月")


def looks_like_month_col(name: str, sample: list[Any] | None = None) -> bool:
    if "月" in (name or ""):
        return True
    vals = sample or []
    return bool(vals) and all(_MONTH_EXTRACT.search(str(v) or "") for v in vals[:8] if v is not None)


def wants_rank_sort(text: str) -> bool:
    t = strip_table_from_prompt(text)
    return bool(re.search(r"(最高|最多|最低|最少|top\s*\d+|前\s*\d+)", t, re.I))


def attach_src_order(df: Any) -> Any:
    if "_src_order" in df.columns:
        return df
    out = df.copy()
    out.insert(0, "_src_order", range(1, len(out) + 1))
    return out


def drop_src_order(
    columns: list[str], rows: list[tuple]
) -> tuple[list[str], list[tuple]]:
    if "_src_order" not in columns:
        return columns, rows
    idx = columns.index("_src_order")
    cols = [c for c in columns if c != "_src_order"]
    out_rows = [tuple(v for i, v in enumerate(row) if i != idx) for row in rows]
    return cols, out_rows


def requested_sheet_stats(text: str) -> dict[str, bool]:
    """用户点名要写入工作簿的统计，不只写在对话里。"""
    hint = strip_table_from_prompt(text)
    blob = hint or (text or "")
    mean = bool(re.search(r"(平均|均值|月均)", blob))
    total = bool(re.search(r"(汇总|合计|总计|总销量|总量|总和)", blob))
    xmax = bool(re.search(r"(最大值|各列最大)", blob))
    xmin = bool(re.search(r"(最小值|各列最小)", blob))
    top = bool(
        re.search(
            r"(销量最高|最高销量|最高的?商品|主销|爆款|卖得最|每个月.{0,8}最高)",
            blob,
        )
    )
    if re.search(r"导出", blob) and not (mean or xmax or xmin or top):
        total = True
    if (mean or top) and not total:
        total = True
    return {"sum": total, "mean": mean, "max": xmax, "min": xmin, "top": top}


def wants_plain_totals(text: str) -> bool:
    hint = strip_table_from_prompt(text)
    if re.search(r"按\s*(类别|品类|分类|产品|项目|科目)", hint):
        return False
    return any(requested_sheet_stats(text).values())


def _stat_label_row(
    *,
    label: str,
    work,
    nums: list[str],
    cat_cols: list[str],
    extra_cols: list[str],
    reducer,
) -> dict:
    row: dict[str, Any] = {}
    for c in work.columns:
        if c in nums or c in extra_cols:
            row[c] = reducer(work[c])
        else:
            row[c] = ""
    if cat_cols:
        row[cat_cols[0]] = label
    return row


def add_row_and_col_totals(
    con,
    table_map: dict[str, str],
    text: str = "",
) -> tuple[list[str], list[tuple], str, list[tuple[str, list[str], list[tuple]]]]:
    """按用户指令写入合计/月均等行，并附「指标」表。"""
    import pandas as pd

    stats = requested_sheet_stats(text)
    tbl = next(iter(table_map.values()))
    names = [c[0] for c in con.execute(f'DESCRIBE "{tbl}"').fetchall()]
    order = 'ORDER BY "_src_order"' if "_src_order" in names else ""
    df = con.execute(f'SELECT * FROM "{tbl}" {order}').fetchdf()
    if "_src_order" in df.columns:
        df = df.drop(columns=["_src_order"])
    types = {c[0]: c[1] for c in con.execute(f'DESCRIBE "{tbl}"').fetchall()}
    nums = [
        c
        for c in df.columns
        if is_numeric_type(str(types.get(c, ""))) and c != "_src_order"
    ]
    work = df.copy()
    extra: list[str] = []
    notes: list[str] = ["已按原表顺序保留明细"]
    if nums and stats.get("sum"):
        work["行合计"] = work[nums].sum(axis=1, numeric_only=True)
        extra.append("行合计")
        notes.append("行合计列")
    if nums and stats.get("top"):
        names_top: list[str] = []
        qtys_top: list[Any] = []
        for _, row in work.iterrows():
            best = max(nums, key=lambda c: float(row[c] or 0))
            names_top.append(str(best))
            qtys_top.append(row[best])
        work["当月最高商品"] = names_top
        work["当月最高销量"] = qtys_top
        notes.append("当月最高商品列")
    cat_cols = [c for c in work.columns if c not in nums and c not in extra]
    data = work.copy()
    extra_rows: list[dict[str, Any]] = []
    if nums and stats.get("sum"):
        extra_rows.append(
            _stat_label_row(
                label="合计",
                work=data,
                nums=nums,
                cat_cols=cat_cols,
                extra_cols=extra,
                reducer=lambda s: float(s.sum()),
            )
        )
        notes.append("合计行")
    if nums and stats.get("mean"):
        extra_rows.append(
            _stat_label_row(
                label="月均",
                work=data,
                nums=nums,
                cat_cols=cat_cols,
                extra_cols=extra,
                reducer=lambda s: round(float(s.mean()), 2),
            )
        )
        notes.append("月均行（写入工作簿）")
    if nums and stats.get("max"):
        extra_rows.append(
            _stat_label_row(
                label="最大",
                work=data,
                nums=nums,
                cat_cols=cat_cols,
                extra_cols=extra,
                reducer=lambda s: float(s.max()),
            )
        )
        notes.append("最大行")
    if nums and stats.get("min"):
        extra_rows.append(
            _stat_label_row(
                label="最小",
                work=data,
                nums=nums,
                cat_cols=cat_cols,
                extra_cols=extra,
                reducer=lambda s: float(s.min()),
            )
        )
        notes.append("最小行")
    if extra_rows and stats.get("top") and nums:
        col_tot = {c: float(data[c].sum()) for c in nums}
        best = max(col_tot, key=col_tot.get)
        extra_rows[0]["当月最高商品"] = best
        extra_rows[0]["当月最高销量"] = col_tot[best]
    if extra_rows:
        work = pd.concat([work, pd.DataFrame(extra_rows)], ignore_index=True)
    cols = list(work.columns)
    rows = [tuple(r) for r in work.itertuples(index=False, name=None)]

    kpi_sheets: list[tuple[str, list[str], list[tuple]]] = []
    if nums and stats.get("mean"):
        n = max(len(data), 1)
        grand = float(data[nums].sum().sum()) if nums else 0.0
        monthly = round(grand / n, 2)
        kpi_rows: list[tuple] = [
            ("总销量", round(grand, 2), f"{n} 个月合计"),
            ("月均销量", monthly, f"总销量/{n}"),
        ]
        for c in nums:
            kpi_rows.append((f"{c}月均", round(float(data[c].mean()), 2), "各月平均"))
        if "当月最高商品" in data.columns:
            kpi_rows.append(("说明", "当月最高商品见结果表", "按行写入"))
        kpi_sheets.append(("指标", ["指标", "数值", "说明"], kpi_rows))
        notes.append("「指标」表写入总销量与月均销量")
    return cols, rows, "；".join(notes), kpi_sheets


def auto_sql(text: str, con, table_map: dict[str, str]) -> str:
    tbl = next(iter(table_map.values()))
    cols = [c[0] for c in con.execute(f'DESCRIBE "{tbl}"').fetchall()]
    types = {c[0]: c[1] for c in con.execute(f'DESCRIBE "{tbl}"').fetchall()}
    nums = [c for c in cols if is_numeric_type(str(types.get(c, ""))) and c != "_src_order"]
    cats = [c for c in cols if c not in nums and c != "_src_order"]
    hint = strip_table_from_prompt(text)
    dim = match_column(hint, cats)
    metric = match_column(hint, nums)
    limit_m = re.search(r"(?:前|top)\s*(\d+)", hint, re.I)
    limit = int(limit_m.group(1)) if limit_m else 200
    limit = max(1, min(limit, 500))
    vis = ", ".join(f'"{c}"' for c in cols if c != "_src_order")
    order_src = 'ORDER BY "_src_order"' if "_src_order" in cols else ""
    group_asked = bool(re.search(r"(分组|按.{0,8}(汇总|合计|统计))", hint))
    if dim and metric and (group_asked or wants_rank_sort(hint)):
        if looks_like_month_col(dim) and not wants_rank_sort(hint):
            order = (
                f'ORDER BY TRY_CAST(regexp_extract(CAST("{dim}" AS VARCHAR), \'(\\d+)\', 1) AS INTEGER) '
                f'NULLS LAST, "{dim}"'
            )
        elif wants_rank_sort(hint):
            order = "ORDER BY total DESC"
        elif "_src_order" in cols:
            order = 'ORDER BY MIN("_src_order")'
        else:
            order = f'ORDER BY "{dim}"'
        return (
            f'SELECT "{dim}", SUM("{metric}") AS total, COUNT(*) AS n '
            f'FROM "{tbl}" GROUP BY "{dim}" {order} LIMIT {limit}'
        )
    return f"SELECT {vis} FROM \"{tbl}\" {order_src} LIMIT {limit}"


def rows_to_markdown(columns: list[str], rows: list[tuple], *, cap: int = 40) -> str:
    if not columns:
        return "（空结果）"
    lines = [
        "| " + " | ".join(str(c) for c in columns) + " |",
        "| " + " | ".join("---" for _ in columns) + " |",
    ]
    for row in rows[:cap]:
        cells = [str(v).replace("|", "\\|") if v is not None else "" for v in row]
        lines.append("| " + " | ".join(cells) + " |")
    if len(rows) > cap:
        lines.append(f"（仅预览前 {cap} 行，共 {len(rows)} 行）")
    return "\n".join(lines)


def _cell_value(val: Any) -> Any:
    """只写入标量，避免把整段字符串/数组拆进多个格子或写成二进制乱码。"""
    if val is None:
        return None
    if isinstance(val, (bytes, bytearray)):
        return val.decode("utf-8", errors="replace")[:500]
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val != val:  # NaN
            return None
        return val
    if hasattr(val, "item") and not isinstance(val, (str, bytes)):
        try:
            return _cell_value(val.item())
        except Exception:  # noqa: BLE001
            return str(val)[:500]
    if isinstance(val, (list, tuple, dict, set)):
        return str(val)[:500]
    text = str(val)
    if len(text) > 500:
        return text[:500]
    return val if isinstance(val, str) else text


def _normalize_sheet_rows(columns: list[str], rows: list) -> tuple[list[str], list[tuple]]:
    cols = [str(c).strip() or f"col_{i}" for i, c in enumerate(columns or [])]
    if not cols:
        return ["列1"], []
    out: list[tuple] = []
    for row in rows or []:
        if isinstance(row, str):
            cells = [row]
        else:
            try:
                cells = list(row)
            except TypeError:
                cells = [row]
        cells = [_cell_value(v) for v in cells[: len(cols)]]
        if len(cells) < len(cols):
            cells.extend([None] * (len(cols) - len(cells)))
        out.append(tuple(cells))
    return cols, out


def add_sum_column(
    con,
    table_map: dict[str, str],
    *,
    name: str = "总销量",
) -> tuple[list[str], list[tuple], str]:
    """在数值列右侧加一列行合计，不额外写合计行、不写指标表。"""
    tbl = next(iter(table_map.values()))
    df = con.execute(f'SELECT * FROM "{tbl}"').fetchdf()
    if "_src_order" in df.columns:
        df = df.drop(columns=["_src_order"])
    types = {c[0]: str(c[1]) for c in con.execute(f'DESCRIBE "{tbl}"').fetchall()}
    skip = {name, "合计", "行合计", "总销量", "总量"}
    nums = [
        c
        for c in df.columns
        if c not in skip and is_numeric_type(types.get(c, ""))
    ]
    if not nums:
        cols, rows = _normalize_sheet_rows(list(df.columns), [tuple(r) for r in df.itertuples(index=False, name=None)])
        return cols, rows, "没有可合计的数值列，已按原表导出"
    label = (name or "总销量").strip() or "总销量"
    work = df.copy()
    work[label] = work[nums].sum(axis=1, numeric_only=True)
    cols, rows = _normalize_sheet_rows(
        list(work.columns),
        [tuple(r) for r in work.itertuples(index=False, name=None)],
    )
    return cols, rows, f"已新增「{label}」列（{'+'.join(nums)}）"


def write_xlsx(
    columns: list[str],
    rows: list[tuple],
    *,
    sheet: str = "数据",
    stem: str = "excel",
) -> tuple[str, str]:
    return write_xlsx_sheets([(sheet, columns, rows)], stem=stem)


_DATE_SUFFIX = {"月", "日", "年", "号", "时", "分", "秒", "周", "季度"}
_SCALE = {"万": 10_000.0, "千": 1_000.0, "亿": 100_000_000.0}
_INSTRUCTION_RE = re.compile(
    r"(按.{0,8}汇总|汇总数据|并计算|月均销量|统计出|"
    r"导出\s*为?\s*excel|生成\s*excel|做成xlsx|画.+图|绘制.+图|下载表格|"
    r"销量最高|最高的?商品|标题设置|加入一列|增加一列|生成柱状|单价|\d+\s*元)",
    re.I,
)


def parse_numeric_cell(val: Any) -> float | None:
    """把「10支」「8块」「1.5万」「12%」收成数字；「1月」保持非数值。"""
    if val is None:
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    s = str(val).strip().replace(",", "").replace("，", "").replace(" ", "")
    if not s or s.lower() in {"-", "—", "na", "nan", "null", "none"}:
        return None
    m = re.match(r"^([+-]?\d+(?:\.\d+)?)(.*)$", s)
    if not m:
        return None
    n = float(m.group(1))
    rest = (m.group(2) or "").strip()
    if rest in _DATE_SUFFIX:
        return None
    if rest in {"%", "％"}:
        return n / 100.0
    if rest in _SCALE:
        return n * _SCALE[rest]
    if rest.startswith("万"):
        return n * 10_000.0
    if rest == "" or re.fullmatch(r"[\u4e00-\u9fffA-Za-z]{1,4}", rest):
        return n
    return None


def coerce_numeric_columns(df: Any) -> Any:
    import pandas as pd

    out = df.copy()
    for col in out.columns:
        series = out[col]
        parsed = series.map(parse_numeric_cell)
        nonempty = series.astype(str).str.strip().replace("", pd.NA).notna()
        n = int(nonempty.sum())
        if n and int(parsed.notna().sum()) / n >= 0.6:
            out[col] = parsed
    return out


def _split_row(line: str) -> list[str]:
    raw = (line or "").strip()
    if not raw:
        return []
    if raw.startswith("|") and raw.endswith("|"):
        cells = [c.strip() for c in raw.strip("|").split("|")]
        return cells
    if "\t" in raw:
        return [c.strip() for c in raw.split("\t")]
    if raw.count(",") >= 2 and raw.count(",") >= raw.count(" "):
        import csv
        from io import StringIO

        rows = list(csv.reader(StringIO(raw)))
        return [c.strip() for c in (rows[0] if rows else [])]
    return [c for c in re.split(r"\s+", raw) if c]


def _is_md_sep(cells: list[str]) -> bool:
    if not cells:
        return False
    return all(re.match(r"^:?-+:?$", c.replace(" ", "")) for c in cells)


def parse_pasted_table(text: str) -> tuple[list[str], list[tuple]] | None:
    """Markdown / CSV / 空格对齐文本表。同一行末尾的「绘制折线图」等指令会剥掉。"""
    t = text or ""
    fence = re.search(r"```(?:csv|tsv|text|markdown)?\s*([\s\S]+?)```", t, re.I)
    blob = fence.group(1).strip() if fence else t
    raw_lines = [ln.rstrip() for ln in blob.splitlines() if ln.strip()]
    candidates: list[list[str]] = []
    for ln in raw_lines:
        if _INSTRUCTION_RE.search(ln) and not ln.strip().startswith("|"):
            if not re.match(r"^\d+\s*月\b", ln.strip()):
                continue
        cells = _split_row(ln)
        if len(cells) < 2 or _is_md_sep(cells):
            continue
        if cells and re.match(r"^(汇总|合计|导出|统计|计算|绘制|标题|加入)", str(cells[0])):
            continue
        if any(re.search(r"\d+\s*元|标题设置|加入一列", str(c)) for c in cells):
            continue
        cut = None
        for i, cell in enumerate(cells):
            if i > 0 and re.search(r"(绘制|折线图|导出为|汇总数据|月均销量)", str(cell)):
                cut = i
                break
        if cut is not None:
            cells = cells[:cut]
        candidates.append(cells)
    if len(candidates) < 2:
        return None
    header_i = 0
    for i, cells in enumerate(candidates):
        joined = "".join(cells)
        if "月份" in joined or "month" in joined.lower():
            header_i = i
            break
    width = len(candidates[header_i])
    if width < 2:
        counts: dict[int, int] = {}
        for cells in candidates:
            counts[len(cells)] = counts.get(len(cells), 0) + 1
        width = max(counts, key=lambda w: (counts[w], w))
    aligned: list[list[str]] = []
    for cells in candidates[header_i:]:
        if len(cells) < width:
            if re.match(r"^\d+\s*月", str(cells[0] or "")) and len(cells) > width:
                aligned.append(cells[:width])
            continue
        aligned.append(cells[:width])
    if len(aligned) < 2:
        return None
    headers = [str(h).strip() or f"col_{i}" for i, h in enumerate(aligned[0])]
    rows: list[tuple] = []
    month_header = looks_like_month_col(headers[0]) if headers else False
    for r in aligned[1:]:
        if len(r) < width:
            continue
        first = str(r[0] or "").strip()
        if month_header and not re.match(r"^(\d+\s*月|[一二三四五六七八九十]+月)", first):
            continue
        rows.append(tuple(r[:width]))
    if not rows:
        return None
    return headers, rows


def wants_sales_amount(text: str) -> bool:
    hint = strip_table_from_prompt(current_user_text(text))
    return bool(re.search(r"总销售金额|销售金额", hint or ""))


def wants_added_column(text: str) -> bool:
    hint = strip_table_from_prompt(current_user_text(text))
    return bool(re.search(r"(加入|增加|加|添加)\s*一列", hint or ""))


def parse_unit_prices(text: str, columns: list[str]) -> dict[str, float]:
    blob = (text or "").replace("，", " ").replace(",", " ")
    cols = [str(c).strip() for c in columns if str(c or "").strip() and str(c).strip() not in {"月份", "月", "_src_order", "总销售金额"}]
    prices: dict[str, float] = {}
    for name in sorted(cols, key=len, reverse=True):
        m = re.search(rf"{re.escape(name)}\s*[:：]?\s*(\d+(?:\.\d+)?)\s*元", blob)
        if m:
            prices[name] = float(m.group(1))
    for m in re.finditer(r"([\u4e00-\u9fffA-Za-z]+)\s*(\d+(?:\.\d+)?)\s*元", blob):
        token, val = m.group(1), float(m.group(2))
        hit = next((c for c in cols if c == token), None)
        if hit is None:
            hit = next((c for c in sorted(cols, key=len, reverse=True) if c and (token in c or c in token)), None)
        if hit and hit not in prices:
            prices[str(hit)] = val
    return prices


def align_prices_to_columns(prices: dict[str, float], columns: list[str]) -> dict[str, float]:
    cols = [str(c).strip() for c in columns if str(c or "").strip()]
    out: dict[str, float] = {}
    for key, val in (prices or {}).items():
        name = str(key or "").strip()
        if not name:
            continue
        if name in cols:
            out[name] = float(val)
            continue
        hit = next((c for c in sorted(cols, key=len, reverse=True) if name in c or c in name), None)
        if hit:
            out[hit] = float(val)
    return out


def coerce_price_map(raw: Any) -> dict[str, float]:
    if not isinstance(raw, dict):
        return {}
    out: dict[str, float] = {}
    for key, val in raw.items():
        name = str(key or "").strip()
        if not name:
            continue
        try:
            out[name] = float(val)
        except (TypeError, ValueError):
            continue
    return out


def apply_unit_price_amount(
    con,
    table_map: dict[str, str],
    text: str,
    extra_prices: dict[str, float] | None = None,
) -> str | None:
    """按单价（如「铅笔1元」）写入「总销售金额」列。"""
    if con is None or not table_map:
        return None
    extra = coerce_price_map(extra_prices)
    if not wants_sales_amount(text) and not extra:
        return None
    import pandas as pd

    tbl = next(iter(table_map.values()))
    df = con.execute(f'SELECT * FROM "{tbl}"').df()
    nums = [c for c in df.columns if c not in {"月份", "月", "_src_order", "总销售金额"}]
    prices = parse_unit_prices(text, nums)
    prices.update(extra)
    prices = align_prices_to_columns(prices, nums)
    if not prices:
        logger.info("unit prices missing, skip 总销售金额")
        return None
    amount = pd.Series(0.0, index=df.index)
    used = 0
    for col, price in prices.items():
        if col not in df.columns:
            continue
        amount = amount + pd.to_numeric(df[col], errors="coerce").fillna(0) * float(price)
        used += 1
    if used == 0:
        return None
    df["总销售金额"] = amount
    tmp = f"_df_{tbl}"
    try:
        con.unregister(tmp)
    except Exception:  # noqa: BLE001
        pass
    con.execute(f'DROP TABLE IF EXISTS "{tbl}"')
    con.register(tmp, df)
    con.execute(f'CREATE TABLE "{tbl}" AS SELECT * FROM {tmp}')
    try:
        con.unregister(tmp)
    except Exception:  # noqa: BLE001
        pass
    logger.info("added 总销售金额 prices=%s", prices)
    return "总销售金额"


def load_frame(df: Any, *, label: str = "数据") -> tuple[Any, dict[str, str]]:
    import duckdb

    df = coerce_numeric_columns(df)
    df.columns = [str(c).strip() or f"col_{i}" for i, c in enumerate(df.columns)]
    df = attach_src_order(df)
    con = duckdb.connect()
    tbl = unique_table(label, set())
    con.register(f"_df_{tbl}", df)
    con.execute(f'CREATE TABLE "{tbl}" AS SELECT * FROM "_df_{tbl}"')
    return con, {label: tbl}


def load_pasted_table(columns: list[str], rows: list[tuple]) -> tuple[Any, dict[str, str]]:
    import pandas as pd

    df = pd.DataFrame([list(r) for r in rows], columns=columns)
    return load_frame(df, label="数据")


def wants_category_rollup(text: str, columns: list[str], numeric_cols: list[str]) -> bool:
    t = text or ""
    named = any(c in {"类别", "品类", "分类", "产品"} for c in columns)
    if named and len(numeric_cols) == 1:
        return False
    if re.search(r"按\s*(类别|品类|分类|产品|项目|科目)", t):
        return len(numeric_cols) >= 2 or named
    return False


def category_rollup_rows(
    con, table_map: dict[str, str]
) -> tuple[list[str], list[tuple], str]:
    tbl = next(iter(table_map.values()))
    if "_src_order" in df.columns:
        df = df.drop(columns=["_src_order"])
    types = {c[0]: c[1] for c in con.execute(f'DESCRIBE "{tbl}"').fetchall()}
    nums = [
        c
        for c in df.columns
        if is_numeric_type(str(types.get(c, ""))) and c != "_src_order"
    ]
    if len(nums) >= 2:
        long = df.melt(value_vars=nums, var_name="类别", value_name="数量")
        grouped = (
            long.groupby("类别", as_index=False)["数量"]
            .sum()
            .sort_values("数量", ascending=False)
        )
        cols = ["类别", "数量"]
        rows = [tuple(r) for r in grouped.itertuples(index=False, name=None)]
        return cols, rows, "已将多列指标按类别（列名）汇总"
    cats = [c for c in df.columns if c not in nums]
    dim = next((c for c in cats if c in {"类别", "品类", "分类", "产品"}), None)
    dim = dim or (cats[0] if cats else None)
    metric = nums[0] if nums else None
    if not dim or not metric:
        return [], [], "没有可按类别汇总的列"
    grouped = (
        df.groupby(dim, as_index=False)[metric]
        .sum()
        .sort_values(metric, ascending=False)
    )
    cols = [dim, metric]
    rows = [tuple(r) for r in grouped.itertuples(index=False, name=None)]
    return cols, rows, f"已按 {dim} 汇总 {metric}"


def write_xlsx_sheets(
    sheets: list[tuple[str, list[str], list[tuple]]],
    *,
    stem: str = "excel",
) -> tuple[str, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=11)
    fill = PatternFill("solid", fgColor="1F4E79")
    first = True
    for title, columns, rows in sheets:
        ws = wb.active if first else wb.create_sheet()
        first = False
        safe_title = re.sub(r"[\\/*?:\[\]]", "_", (title or "数据"))[:31] or "数据"
        ws.title = safe_title
        columns, rows = _normalize_sheet_rows(list(columns or []), list(rows or []))
        for i, col in enumerate(columns, 1):
            cell = ws.cell(1, i, _cell_value(col))
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                cell = ws.cell(r, c, _cell_value(val))
                cell.font = body_font
        for i, col in enumerate(columns, 1):
            width = min(40, max(10, len(str(col)) + 4))
            ws.column_dimensions[get_column_letter(i)].width = width
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    name = f"{re.sub(r'[^\w.\-一-龥]+', '_', stem)[:40]}-{uuid.uuid4().hex[:8]}.xlsx"
    path = GENERATED_DIR / name
    wb.save(path)
    return _public_url(name), name


_KIND_WORDS: list[tuple[str, tuple[str, ...]]] = [
    ("pie", ("饼状图", "饼图", "pie chart")),
    ("line", ("折线图", "走势图", "趋势图", "曲线图", "line chart", "折线")),
    ("bar", ("柱状图", "柱形图", "条形图", "bar chart", "柱状", "柱形", "条形")),
    ("scatter", ("散点图", "scatter")),
    ("hist", ("直方图", "直方", "hist")),
    ("radar", ("雷达图", "雷达")),
]

_KIND_RE = {
    "line": r"折线|走势|趋势|曲线|line\s*chart",
    "bar": r"柱状|柱形|条形|bar\s*chart",
    "pie": r"饼|占比|pie",
}

_EACH_SERIES_RE = re.compile(
    r"(?:每(?:一种|一种|种|个|类)|各(?:个|种|类)?|所有|全部).{0,20}"
    r"(?:商品|品类|产品|系列|指标|列|项)"
    r"|each\s+(?:product|series|category|column)"
    r"|one\s+chart\s+per\s+(?:product|series)",
    re.I,
)
_MONTH_FACET_RE = re.compile(
    r"每个月|每月|各月|按月|分月|逐月|月份各|每月一|逐个月份"
    r"|per\s*month|each\s*month",
    re.I,
)
_MULTI_IMG_RE = re.compile(
    r"多[个张幅]|若干[张幅个]?|好几[张个]|分别|分开|单独|各自|"
    r"各一[张幅个]|分开展示|分开画|分别画|一张一[个张]|"
    r"separately|multiple\s+(?:charts?|plots?|figures?)",
    re.I,
)
_COMBO_RE = re.compile(
    r"只要一[张幅个]图|不要[两多]张|同一[张幅]图|一张图(?:里|中|内)|"
    r"图像中还要|同[一张幅]图|组合图|双轴|叠加(?:在一起|显示)|"
    r"同时(?!再).{0,30}(?:折线|柱状)|(?:柱状|折线).{0,24}同时(?!再)|"
    r"还要有.{0,16}(?:折线|柱状)|combo\s*chart|overlay",
    re.I,
)
_SEPARATE_CHART_RE = re.compile(
    r"再生成一个|另外(?:再)?(?:生成|画|做)|还要一[张幅个]|再画一|"
    r"同时再(?:生成|画|做)|分别(?:生成|画)|各一[张幅个]"
)
_COUNT_CHARTS_RE = re.compile(r"[两二三四五六七八九十\d]+[张个幅条](?:图|折线|柱|饼)?")


def wants_combo_chart(text: str) -> bool:
    """柱状+折线画在同一张图里（组合图），而不是两张图。"""
    hint = strip_table_from_prompt(text)
    if _SEPARATE_CHART_RE.search(hint or ""):
        return False
    has_bar = bool(re.search(_KIND_RE["bar"], hint, re.I))
    has_line = bool(re.search(_KIND_RE["line"], hint, re.I))
    if not (has_bar and has_line):
        return False
    if re.search(r"只要一[张幅个]图|同一[张幅]|一张图(?:里|中|内)|不要[两多]张", hint):
        return True
    return bool(_COMBO_RE.search(hint))


def wants_amount_line_chart(text: str) -> bool:
    hint = strip_table_from_prompt(current_user_text(text))
    return bool(re.search(r"总销售金额.{0,16}折线|折线.{0,12}总销售金额", hint or ""))


def wants_combo_total_line(text: str) -> bool:
    """柱状看各品类，折线只画总销量/合计变化（不是每类一条折线）。"""
    hint = strip_table_from_prompt(text)
    if wants_amount_line_chart(hint) or _SEPARATE_CHART_RE.search(hint or ""):
        return False
    if wants_total_as_extra_series(hint) and not re.search(r"折线|line", hint, re.I):
        return False
    if re.search(
        r"(总销量|合计|总体).{0,12}(变化)?折线|(变化)?折线.{0,8}(总销量|合计)",
        hint,
    ):
        return True
    has_bar = bool(re.search(_KIND_RE["bar"], hint, re.I))
    has_line = bool(re.search(_KIND_RE["line"], hint, re.I))
    has_total = bool(re.search(r"总销量|合计|总体|总量", hint))
    each_line = bool(re.search(r"各(?:类|种)(?:商品|品类).{0,12}折线|每种.{0,8}折线", hint))
    return bool(has_bar and has_line and has_total and not each_line)


def combo_line_mode(text: str) -> str:
    """组合图折线：each=各类商品；total=仅当用户明确要合计/总体。"""
    hint = strip_table_from_prompt(text)
    if wants_combo_total_line(hint):
        return "total"
    each = bool(
        re.search(
            r"各(?:类|种|个)?(?:商品|品类)|每种|每个商品|分品类|各产品",
            hint,
        )
    )
    total = bool(re.search(r"合计|总体|总销量|总量|总商品", hint))
    if each:
        return "each"
    if total:
        return "total"
    return "each"


def correct_chart_spec(spec: dict[str, Any] | None, user_text: str) -> dict[str, Any]:
    """模型提案 + 用户原话终审。缺 line_mode 时不用静默 each。"""
    spec = copy.deepcopy(spec or {})
    hint = strip_table_from_prompt(user_text)
    jobs = spec.get("charts")
    targets: list[dict[str, Any]]
    if isinstance(jobs, list) and jobs:
        targets = [j for j in jobs if isinstance(j, dict)]
    else:
        targets = [spec]
    extra_col = wants_total_as_extra_series(hint)
    combo_total = wants_combo_total_line(hint)
    total_only_intent = (
        _is_total_ctx(hint) and not extra_col and not combo_total
    )
    if extra_col:
        spec["include_total_as_series"] = True
        spec["total_only"] = False
        for j in targets:
            j["include_total_as_series"] = True
            j["total_only"] = False
    if combo_total:
        spec["total_only"] = False
        spec["line_mode"] = "total"
        for j in targets:
            if str(j.get("kind") or "") == "pie":
                continue
            j["kind"] = "combo"
            j["line_mode"] = "total"
            j["total_only"] = False
        if targets is spec or spec in targets:
            spec["kind"] = "combo"
            spec["line_mode"] = "total"
    elif total_only_intent:
        spec["total_only"] = True
        for j in targets:
            if str(j.get("kind") or spec.get("kind") or "bar") in {"bar", "line", ""}:
                j["total_only"] = True
    for j in targets:
        kind = str(j.get("kind") or spec.get("kind") or "")
        if kind != "combo":
            continue
        lm = j.get("line_mode") or spec.get("line_mode")
        if lm not in {"each", "total"}:
            j["line_mode"] = combo_line_mode(hint)
        if combo_total:
            j["line_mode"] = "total"
    return spec


def wants_chart_output(text: str) -> bool:
    hint = strip_table_from_prompt(current_user_text(text))
    return bool(
        re.search(
            r"可视化|柱状|柱形|条形|折线|饼图|饼状|散点|直方|雷达|"
            r"chart|plot|出图|画图|做[一张幅]图",
            hint,
            re.I,
        )
    )


def wants_workbook_output(text: str) -> bool:
    hint = strip_table_from_prompt(current_user_text(text))
    return bool(
        re.search(
            r"xlsx|\.xls\b|xls表格|生成\s*xls|做成\s*xls|"
            r"excel|工作簿|表格文件|生成表|做成表|导出|下载表|给我表",
            hint,
            re.I,
        )
    )


def wants_pie_chart(text: str) -> bool:
    hint = strip_table_from_prompt(current_user_text(text))
    return bool(re.search(r"饼图|饼|pie|占比", hint, re.I))


def expand_chart_spec(spec: dict[str, Any] | None, user_text: str) -> dict[str, Any]:
    """校正参数后，把用户点名但 spec 漏掉的图种补进 charts。"""
    spec = correct_chart_spec(spec, user_text)
    if not wants_chart_output(user_text):
        return spec
    raw = spec.get("charts")
    if isinstance(raw, list) and raw:
        jobs = [dict(j) for j in raw if isinstance(j, dict)]
    else:
        jobs = [{k: spec[k] for k in spec if k != "charts"}]
    kinds = [str(j.get("kind") or "").lower() for j in jobs]
    hint = strip_table_from_prompt(user_text)
    if wants_combo_total_line(hint):
        if "combo" not in kinds:
            jobs.insert(0, {"kind": "combo", "line_mode": "total"})
        else:
            for j in jobs:
                if str(j.get("kind") or "").lower() == "combo":
                    j["line_mode"] = "total"
    if wants_pie_chart(hint) and "pie" not in [str(j.get("kind") or "").lower() for j in jobs]:
        jobs.append({"kind": "pie", "title": "销量占比"})
    if len(jobs) == 1:
        merged = dict(spec)
        merged.update(jobs[0])
        merged.pop("charts", None)
        return merged
    shared = {
        k: v
        for k, v in spec.items()
        if k not in {"kind", "line_mode", "charts", "title", "split", "facet"}
    }
    return {**shared, "charts": jobs}


def _find_chart_markers(hint: str) -> list[tuple[int, int, str]]:
    ranked: list[tuple[int, str, str]] = []
    for kind, toks in _KIND_WORDS:
        for tok in toks:
            ranked.append((len(tok), tok, kind))
    ranked.sort(key=lambda x: -x[0])
    taken = [False] * (len(hint) + 1)
    found: list[tuple[int, int, str]] = []
    for _, tok, kind in ranked:
        start = 0
        while True:
            i = hint.find(tok, start)
            if i < 0:
                break
            if not any(taken[i : i + len(tok)]):
                for j in range(i, i + len(tok)):
                    taken[j] = True
                found.append((i, i + len(tok), kind))
            start = i + 1
    found.sort(key=lambda x: x[0])
    return found


def extract_chart_title(text: str) -> str | None:
    hint = strip_table_from_prompt(current_user_text(text))
    m = re.search(
        r"(?:把)?标题\s*(?:设置?为|设置成|设为|设成|改成|改为|用|是|为)\s*[：:=]?\s*"
        r"[「『\"“']?([^」』\"”'\n，,。；;]{2,40})",
        hint or "",
        re.I,
    )
    if m:
        title = m.group(1).strip()
        title = re.sub(r"(并)?(生成|加入|绘制|画).*$", "", title).strip(" ：:=的")
        if len(title) >= 2:
            return title[:40]
    m = re.search(r"[「『]([^」』]{2,40})[」』]", hint or "")
    if m:
        return m.group(1).strip()[:40]
    return None


_SALES_PLACEHOLDER_TITLES = {
    "销量情况",
    "销量情况（含总销量）",
    "销量情况(含总销量)",
    "销量柱状图 + 合计变化折线",
    "销量柱状图 + 各类商品变化折线",
    "销量占比",
    "总体销量占比",
    "总体销量（各月合计）",
    "总体销量(各月合计)",
}
_SCORE_HINT_RE = re.compile(
    r"recall|f1|bleu|rouge|score|acc(uracy)?|precision|得分|指标|评测|模型",
    re.I,
)
_SALES_HINT_RE = re.compile(r"销量|销售额|文具|商品")


def _looks_like_sales_placeholder(title: str | None, user_text: str) -> bool:
    t = (title or "").strip()
    if t not in _SALES_PLACEHOLDER_TITLES:
        return False
    blob = strip_table_from_prompt(current_user_text(user_text))
    return not bool(_SALES_HINT_RE.search(blob or ""))


def infer_chart_title(
    text: str,
    *,
    series: list[str] | None = None,
    kind: str = "bar",
) -> str | None:
    """未指定标题时，按用户原句和列名生成图题。"""
    explicit = extract_chart_title(text)
    if explicit:
        return explicit
    hint = strip_table_from_prompt(current_user_text(text))
    m = re.search(
        r"(?:绘制|画|生成|做)[一张幅个]?(?P<topic>.{1,24}?)(?P<ckind>柱状图|柱形图|条形图|折线图|饼图|组合图)",
        hint or "",
    )
    if m:
        topic = m.group("topic").strip(" 的了下")
        if topic:
            return f"{topic}{m.group('ckind')}"[:40]
    cols = " ".join(str(c) for c in (series or []) if c and c != "合计")
    blob = f"{hint or ''} {cols}"
    if _SCORE_HINT_RE.search(blob):
        if kind == "pie":
            return "指标占比"
        return "模型得分对比"
    names = [str(c) for c in (series or []) if c and c != "合计"]
    if names and not _SALES_HINT_RE.search(blob):
        if len(names) == 1:
            return str(names[0])
        return ("、".join(names[:3]) + " 对比")[:40]
    return None


def infer_y_label(text: str, series: list[str] | None = None) -> str:
    cols = " ".join(str(c) for c in (series or []) if c)
    blob = f"{strip_table_from_prompt(current_user_text(text)) or ''} {cols}"
    if series and any("金额" in str(c) for c in series):
        return "金额"
    if _SCORE_HINT_RE.search(blob):
        return "得分"
    if re.search(r"占比|比例|percent", blob, re.I):
        return "占比"
    if _SALES_HINT_RE.search(blob):
        return "销量"
    names = [str(c) for c in (series or []) if c and c != "合计"]
    if len(names) == 1:
        return names[0]
    return "数值"


def resolve_chart_title(
    user_text: str,
    spec_title: str | None,
    *,
    series: list[str] | None = None,
    kind: str = "bar",
) -> str | None:
    explicit = extract_chart_title(user_text)
    if explicit:
        if kind == "line" and series and any("金额" in str(c) for c in series) and "金额" not in explicit:
            pass
        else:
            return explicit
    spec = (spec_title or "").strip() or None
    if spec and _looks_like_sales_placeholder(spec, user_text):
        spec = None
    return spec or infer_chart_title(user_text, series=series, kind=kind)


def _is_monthly_facet(ctx: str) -> bool:
    return bool(_MONTH_FACET_RE.search(ctx or ""))


def wants_total_as_extra_series(ctx: str) -> bool:
    """在各类商品之外再加合计列（不是只画总体）。"""
    t = (ctx or "").replace(" ", "")
    if "增加一列总销量" in t or "加一列总销量" in t:
        return True
    return bool(
        re.search(
            r"(增加|加|加上|再加|添加).{0,10}(一列|一栏).{0,12}(总销量|合计|总量|总和|总计)"
            r"|(总销量|合计).{0,6}(一列|一栏)",
            ctx or "",
        )
    )


def _is_total_ctx(ctx: str) -> bool:
    if wants_total_as_extra_series(ctx):
        return False
    return any(k in (ctx or "") for k in ("总体", "总销量", "合计", "全部销量", "总量", "总和"))


def _multi_of_kind(hint: str, kind: str) -> bool:
    pat = _KIND_RE.get(kind)
    if not pat:
        return False
    blob = hint or ""
    return bool(
        re.search(
            rf"(?:多[个张幅]|若干|好几[张个]|分别|分开|[两二三四五六七八九十\d]+[张个幅条]).{{0,20}}(?:{pat})"
            rf"|(?:{pat}).{{0,20}}(?:多[个张幅]|若干|分别|分开|各一|[两二三四五六七八九十\d]+[张个幅])",
            blob,
            re.I,
        )
    )


def _should_split_xy(kind: str, ctx: str, hint: str) -> bool:
    """折线/柱状是否按系列拆成多张。看「每种/多个/两张」等意图，不绑死某句原话。"""
    if kind not in {"bar", "line"}:
        return False
    if _is_total_ctx(ctx) and not _EACH_SERIES_RE.search(ctx or ""):
        return False
    blob = f"{ctx or ''} {hint or ''}"
    if _EACH_SERIES_RE.search(blob):
        return True
    if _multi_of_kind(ctx or "", kind) or _multi_of_kind(hint or "", kind):
        return True
    if _MULTI_IMG_RE.search(ctx or "") and re.search(_KIND_RE[kind], blob, re.I):
        return True
    if _COUNT_CHARTS_RE.search(ctx or "") and re.search(_KIND_RE[kind], blob, re.I):
        return True
    return False


def parse_chart_jobs(text: str, nums: list[str] | None = None) -> list[dict[str, Any]]:
    """一句话里多个图种拆成多个任务：折线图 + 总体柱状图 → 两份。"""
    hint = strip_table_from_prompt(text)
    marks = _find_chart_markers(hint)
    if not marks:
        kind = _chart_kind_fallback(text)
        return [
            {
                "kind": kind,
                "ctx": hint,
                "split": _should_split_xy(kind, hint, hint),
            }
        ]
    jobs: list[dict[str, Any]] = []
    prev = 0
    for start, end, kind in marks:
        ctx = hint[prev:start]
        jobs.append({"kind": kind, "ctx": ctx, "split": False})
        prev = end
    if not jobs:
        return [{"kind": "bar", "ctx": hint, "split": False}]
    tail = hint[prev:]
    if tail.strip():
        jobs[-1]["ctx"] = str(jobs[-1]["ctx"]) + tail
    for j in jobs:
        ctx = str(j.get("ctx") or "")
        kind = str(j.get("kind") or "")
        j["split"] = _should_split_xy(kind, ctx, hint)
    if wants_combo_chart(text):
        kinds = {str(j.get("kind")) for j in jobs}
        if "bar" in kinds and "line" in kinds:
            pies = [j for j in jobs if j.get("kind") == "pie"]
            combo = {"kind": "combo", "ctx": hint, "split": False}
            jobs = [combo] + pies
    return jobs


def resolve_job_series(job: dict[str, Any], nums: list[str]) -> tuple[list[str], bool]:
    ctx = str(job.get("ctx") or "")
    if job.get("kind") in {"bar", "line"} and _is_total_ctx(ctx):
        return ["合计"], True
    named = [c for c in nums if c and str(c) in ctx]
    if named:
        return named, False
    return list(nums), False


def _chart_kind_fallback(text: str) -> str:
    hint = strip_table_from_prompt(text)
    blob = f"{text or ''} {hint}"
    if any(k in blob for k in ("饼图", "饼", "pie", "占比")):
        return "pie"
    if any(k in blob for k in ("柱状", "柱形", "条形", "bar chart")):
        return "bar"
    if any(k in blob for k in ("折线", "走势图", "趋势图", "line chart")):
        return "line"
    if any(k in blob for k in ("散点", "scatter")):
        return "scatter"
    if any(k in blob for k in ("直方", "hist")):
        return "hist"
    if any(k in blob for k in ("雷达", "radar")):
        return "radar"
    return "bar"


def chart_kind(text: str) -> str:
    jobs = parse_chart_jobs(text)
    return str(jobs[0]["kind"]) if jobs else "bar"


def wants_multi_pie(text: str) -> bool:
    hint = strip_table_from_prompt(text)
    return _is_monthly_facet(hint) or _multi_of_kind(hint, "pie")


def wants_split_series(text: str) -> bool:
    hint = strip_table_from_prompt(text)
    return _should_split_xy("line", hint, hint) or _should_split_xy("bar", hint, hint)


def _cjk_fontprops():
    from matplotlib.font_manager import FontProperties

    if _CJK_FONT.is_file():
        return FontProperties(fname=str(_CJK_FONT))
    return None


def _apply_title(ax, text: str | None) -> None:
    t = (text or "").strip()
    if not t:
        return
    fp = _cjk_fontprops()
    kwargs: dict = {"fontsize": 13, "pad": 10}
    if fp is not None:
        kwargs["fontproperties"] = fp
    ax.set_title(t, **kwargs)


def _place_legend(fig, ax, n_series: int) -> None:
    if n_series <= 1:
        return
    fig.subplots_adjust(right=0.76)
    ax.legend(
        loc="center left",
        bbox_to_anchor=(1.02, 0.5),
        borderaxespad=0.0,
        frameon=True,
        fontsize=8,
        framealpha=0.95,
    )


def _set_category_ticks(ax, labels: list[str], positions=None) -> None:
    labs = [str(x) for x in labels]
    if positions is None:
        ax.set_xticklabels(labs)
    else:
        ax.set_xticks(positions)
        ax.set_xticklabels(labs)
    longest = max((len(s) for s in labs), default=0)
    if longest >= 8:
        for tick in ax.get_xticklabels():
            tick.set_rotation(22)
            tick.set_ha("right")


def _new_axes(figsize=(8.4, 4.6)):
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=figsize)
    fig.set_facecolor("white")
    ax.set_facecolor("white")
    fig.subplots_adjust(left=0.11, right=0.97, top=0.88, bottom=0.22)
    return fig, ax


def _save_chart_png(fig, stem: str) -> tuple[str, str]:
    import matplotlib.pyplot as plt

    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    name = f"{re.sub(r'[^\w.\-一-龥]+', '_', stem)[:40]}-{uuid.uuid4().hex[:8]}.png"
    path = GENERATED_DIR / name
    fig.savefig(path, dpi=140, bbox_inches="tight", pad_inches=0.4, facecolor="white")
    plt.close(fig)
    return _public_url(name), name


def _draw_pie(labels: list[str], values: list[float], title: str):
    fig, ax = _new_axes((7.2, 6.4))
    ax.pie(values, labels=labels, autopct="%1.1f%%")
    _apply_title(ax, title)
    return fig


def _pies_from_wide_df(
    df,
    series: list[str],
    stem: str,
    *,
    facet: bool,
    include_overall: bool = True,
    overall_title: str = "总体销量占比",
) -> tuple[list[str], list[str], list[str]]:
    urls: list[str] = []
    names: list[str] = []
    mds: list[str] = []
    pie_labels = [str(c) for c in series]
    if facet:
        for rec in df.itertuples(index=False):
            row_label = str(rec[0])
            vals = [float(v or 0) for v in rec[1 : 1 + len(series)]]
            fig = _draw_pie(pie_labels, vals, f"{row_label} 各商品销量占比")
            u, n = _save_chart_png(fig, f"{stem}-{row_label}")
            urls.append(u)
            names.append(n)
            mds.append(
                f"### {row_label}\n"
                + rows_to_markdown(["品类", "销量"], list(zip(pie_labels, vals)))
            )
    if include_overall or not facet:
        totals = [float(df[c].sum()) for c in series]
        fig = _draw_pie(pie_labels, totals, overall_title)
        u, n = _save_chart_png(fig, f"{stem}-整体")
        urls.append(u)
        names.append(n)
        mds.append(
            "### 整体\n" + rows_to_markdown(["品类", "销量"], list(zip(pie_labels, totals)))
        )
    return urls, names, mds


def _fit_xy(ax, xmin: float, xmax: float, *, n_ticks: int = 0) -> None:
    """给左右数据点留白，并保证底部刻度可见。"""
    span = max(xmax - xmin, 1.0)
    pad = max(span * 0.12, 0.55)
    ax.set_xlim(xmin - pad, xmax + pad)
    lo, hi = ax.get_ylim()
    if lo >= 0:
        ax.set_ylim(0, hi * 1.16 if hi else 1)
    else:
        ax.set_ylim(lo * 1.08, hi * 1.12)
    ax.tick_params(axis="x", labelsize=10, pad=6)
    ax.tick_params(axis="y", labelsize=10)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ax.set_axisbelow(True)


def _plot_bar_or_line(
    kind: str,
    labels: list[str],
    df,
    series: list[str],
    x_name: str,
    *,
    title: str | None = None,
    ylabel: str | None = None,
):
    import numpy as np

    fig, ax = _new_axes((8.6, 4.6))
    xs = list(range(len(labels)))
    palette = ["#1F4E79", "#C45911", "#548235", "#7030A0", "#C00000", "#00A3A1"]
    requested = (title or "").strip() or None
    if kind == "line":
        for i, col in enumerate(series):
            ax.plot(
                xs,
                df[col].tolist(),
                marker="o",
                color=palette[i % len(palette)],
                label=str(col),
            )
        ax.set_xticks(xs)
        _set_category_ticks(ax, labels)
        fallback = (
            f"{x_name} 折线（{len(series)} 项）"
            if len(series) > 1
            else f"{series[0]} 变化"
        )
        ax.set_ylabel(ylabel or (str(series[0]) if len(series) == 1 else "数值"))
        _place_legend(fig, ax, len(series))
        _fit_xy(ax, 0, max(len(labels) - 1, 0), n_ticks=len(labels))
        _apply_title(ax, requested or fallback)
    else:
        if len(series) == 1:
            ax.bar(xs, df[series[0]].tolist(), color="#1F4E79", width=0.55)
            ax.set_xticks(xs)
            fallback = f"{series[0]} 按 {x_name}"
            xmax = float(len(labels) - 1)
            _set_category_ticks(ax, labels)
        else:
            idx = np.arange(len(labels), dtype=float)
            width = min(0.8 / len(series), 0.22)
            for i, col in enumerate(series):
                ax.bar(
                    idx + i * width,
                    df[col].tolist(),
                    width * 0.92,
                    label=str(col),
                    color=palette[i % len(palette)],
                )
            ax.set_xticks(idx + width * (len(series) - 1) / 2)
            fallback = f"{x_name} 对比（{len(series)} 项）"
            xmax = float(idx[-1] + width * (len(series) - 1)) if len(labels) else 1.0
            _set_category_ticks(ax, labels, ax.get_xticks())
            _place_legend(fig, ax, len(series))
        ax.set_ylabel(ylabel or "数值")
        _fit_xy(ax, -0.2, xmax, n_ticks=len(labels))
        _apply_title(ax, requested or fallback)
    ax.set_xlabel(str(x_name))
    return fig


def _plot_combo(
    labels: list[str],
    df,
    series: list[str],
    x_name: str,
    *,
    title: str | None = None,
    line_mode: str = "each",
    ylabel: str | None = None,
):
    """柱状看各品类销量；折线默认各类商品变化，仅在 line_mode=total 时画合计。"""
    import numpy as np

    fig, ax = _new_axes((8.2, 4.8))
    n = max(len(series), 1)
    idx = np.arange(len(labels), dtype=float)
    width = min(0.8 / n, 0.22)
    palette = ["#1F4E79", "#C45911", "#548235", "#7030A0", "#C00000", "#00A3A1"]
    mid = idx + width * (n - 1) / 2
    for i, col in enumerate(series):
        color = palette[i % len(palette)]
        ax.bar(
            idx + i * width,
            df[col].tolist(),
            width * 0.9,
            label=str(col),
            color=color,
            alpha=0.88,
            zorder=2,
        )
        if line_mode != "total":
            ax.plot(
                mid,
                df[col].tolist(),
                marker="o",
                color=color,
                linewidth=1.8,
                zorder=3,
            )
    ax.set_xticks(mid)
    _set_category_ticks(ax, labels, mid)
    ax.set_ylabel(ylabel or "数值")
    ax.set_xlabel(str(x_name))
    default_title = (
        "柱状图 + 合计变化折线"
        if line_mode == "total"
        else "柱状图 + 各系列变化折线"
    )
    _apply_title(ax, title or default_title)
    xmax = float(idx[-1] + width * (n - 1)) if len(labels) else 1.0
    _fit_xy(ax, -0.2, xmax, n_ticks=len(labels))
    if line_mode == "total":
        totals = df[series].sum(axis=1).astype(float).tolist()
        ax2 = ax.twinx()
        ax2.plot(mid, totals, color="#C00000", marker="o", linewidth=2.2, label="合计变化", zorder=4)
        ax2.set_ylabel("合计")
        tmax = max(totals) if totals else 1.0
        ax2.set_ylim(0, tmax * 1.18 if tmax else 1)
        h1, l1 = ax.get_legend_handles_labels()
        h2, l2 = ax2.get_legend_handles_labels()
        fig.subplots_adjust(right=0.70)
        ax.legend(
            h1 + h2,
            l1 + l2,
            loc="center left",
            bbox_to_anchor=(1.12, 0.5),
            frameon=True,
            fontsize=8,
        )
    else:
        _place_legend(fig, ax, len(series))
    return fig


def _pick_chart_axes(
    text: str, cats: list[str], nums: list[str]
) -> tuple[str | None, list[str]]:
    hint = strip_table_from_prompt(text)
    month = next((c for c in cats if looks_like_month_col(c)), None)
    x = match_column(hint, cats) or month or (cats[0] if cats else None)
    if x == "_src_order":
        x = month or next((c for c in cats if c != "_src_order"), None)
    named = [c for c in nums if c and str(c) in hint]
    if named:
        ys = named
    else:
        ys = list(nums)
    return x, ys


def _setup_matplotlib() -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    if _CJK_FONT.is_file():
        try:
            from matplotlib import font_manager

            font_manager.fontManager.addfont(str(_CJK_FONT))
            plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial"]
        except Exception:  # noqa: BLE001
            plt.rcParams["font.sans-serif"] = ["SimHei", "Arial"]
    plt.rcParams["axes.unicode_minus"] = False


def _match_series_names(requested: list[str] | None, nums: list[str]) -> list[str]:
    if not requested:
        return list(nums)
    out: list[str] = []
    for raw in requested:
        name = str(raw or "").strip()
        if not name or name in {"合计", "总销量", "总量"}:
            continue
        if name in nums:
            if name not in out:
                out.append(name)
            continue
        hit = next((c for c in nums if name in str(c) or str(c) in name), None)
        if hit and hit not in out:
            out.append(hit)
    return out or list(nums)


def _fetch_wide_df(con, table_map: dict[str, str], x: str, nums: list[str]):
    tbl = next(iter(table_map.values()))
    names = [c[0] for c in con.execute(f'DESCRIBE "{tbl}"').fetchall()]
    has_order = "_src_order" in names
    order_sql = 'ORDER BY "_src_order"' if has_order else (
        f'ORDER BY TRY_CAST(regexp_extract(CAST("{x}" AS VARCHAR), \'(\\d+)\', 1) AS INTEGER) '
        f'NULLS LAST, "{x}"'
        if looks_like_month_col(x)
        else f'ORDER BY "{x}"'
    )
    quoted = ", ".join(f'"{c}"' for c in nums)
    return con.execute(
        f'SELECT "{x}" AS label, {quoted} FROM "{tbl}" '
        f'WHERE "{x}" IS NOT NULL {order_sql}'
    ).fetchdf()


def render_from_spec(
    con,
    table_map: dict[str, str],
    spec: dict[str, Any],
    *,
    stem: str = "chart",
    user_text: str = "",
) -> tuple[str, str, str, list[str]]:
    """按结构化 spec 出图。返回 (url, filename, note, all_urls)。"""
    spec = expand_chart_spec(spec, user_text)
    _setup_matplotlib()
    tbl = next(iter(table_map.values()))
    cols = con.execute(f'DESCRIBE "{tbl}"').fetchall()
    names = [c[0] for c in cols]
    types = {c[0]: c[1] for c in cols}
    nums = [c for c in names if is_numeric_type(str(types.get(c, ""))) and c != "_src_order"]
    kind_hint = str(spec.get("kind") or "").lower()
    series_hint = spec.get("series") if isinstance(spec.get("series"), list) else []
    keep_amount = kind_hint == "line" or any("金额" in str(s) for s in series_hint)
    if wants_sales_amount(user_text) and not keep_amount:
        qty = [c for c in nums if c not in {"总销售金额", "销售金额", "销售额"}]
        if qty:
            nums = qty
    cats = [c for c in names if c not in nums and c != "_src_order"]
    month = next((c for c in cats if looks_like_month_col(c)), None)
    raw_charts = spec.get("charts")
    charts: list[dict[str, Any]]
    if isinstance(raw_charts, list) and raw_charts:
        charts = [c for c in raw_charts if isinstance(c, dict)]
    else:
        charts = [spec]
    if not charts:
        raise ValueError("图表 spec 为空")
    x = str(spec.get("x") or charts[0].get("x") or month or (cats[0] if cats else "") or "")
    if not x:
        raise ValueError("柱状/折线图需要类别列或日期列（如月份）")
    if not nums:
        raise ValueError("没有可绘制的数值列")
    df = _fetch_wide_df(con, table_map, x, nums)
    labels = df["label"].astype(str).tolist()
    urls: list[str] = []
    names_out: list[str] = []
    md_chunks: list[str] = []
    for ji, job in enumerate(charts):
        kind = str(job.get("kind") or spec.get("kind") or "bar").lower()
        if kind in {"column", "柱状", "柱形", "条形"}:
            kind = "bar"
        elif kind in {"折线"}:
            kind = "line"
        elif kind in {"饼", "饼图"}:
            kind = "pie"
        series = _match_series_names(
            job.get("series") if isinstance(job.get("series"), list) else spec.get("series"),
            nums,
        )
        include_total = bool(
            job.get("include_total_as_series")
            or spec.get("include_total_as_series")
        )
        total_only = bool(job.get("total_only") or spec.get("total_only"))
        split = bool(job.get("split") or spec.get("split"))
        spec_title = str(job.get("title") or spec.get("title") or "").strip() or None
        title = resolve_chart_title(
            user_text, spec_title, series=series, kind=kind
        )
        y_label = infer_y_label(user_text, series)
        plot_df = df
        if include_total or total_only:
            plot_df = df.copy()
            plot_df["合计"] = plot_df[nums].sum(axis=1)
        if total_only and not include_total:
            series = ["合计"]
            title = title or resolve_chart_title(user_text, "总体销量（各月合计）", series=series, kind=kind)
        elif include_total:
            series = [c for c in series if c != "合计"] or list(nums)
            if "合计" not in series:
                series = list(series) + ["合计"]
            title = title or resolve_chart_title(
                user_text, "销量情况（含总销量）", series=series, kind=kind
            )
        if not series:
            continue
        if kind == "pie":
            facet = bool(job.get("facet") or job.get("facet_pie"))
            include_overall = job.get("include_overall")
            if include_overall is None:
                include_overall = True
            pu, pn, pmd = _pies_from_wide_df(
                plot_df[["label", *[c for c in series if c in plot_df.columns]]],
                [c for c in series if c != "合计"],
                stem,
                facet=facet,
                include_overall=bool(include_overall),
                overall_title=title or infer_chart_title(user_text, series=series, kind="pie") or "占比",
            )
            urls.extend(pu)
            names_out.extend(pn)
            md_chunks.extend(pmd)
            continue
        if kind == "combo":
            line_mode = str(job.get("line_mode") or spec.get("line_mode") or "")
            if line_mode not in {"each", "total"}:
                line_mode = combo_line_mode(user_text) if user_text else "each"
            one = _plot_combo(
                labels,
                plot_df,
                [c for c in series if c != "合计"] or series,
                x,
                title=title,
                line_mode=line_mode if line_mode in {"each", "total"} else "each",
                ylabel=y_label,
            )
            u, n = _save_chart_png(one, f"{stem}-combo")
            urls.append(u)
            names_out.append(n)
        elif split and kind in {"bar", "line"} and len([c for c in series if c != "合计"]) > 1:
            for col in series:
                one = _plot_bar_or_line(kind, labels, plot_df, [col], x, ylabel=y_label)
                u, n = _save_chart_png(one, f"{stem}-{col}")
                urls.append(u)
                names_out.append(n)
        elif kind in {"bar", "line"}:
            one = _plot_bar_or_line(kind, labels, plot_df, series, x, title=title, ylabel=y_label)
            u, n = _save_chart_png(one, f"{stem}-{kind}-{ji}")
            urls.append(u)
            names_out.append(n)
        else:
            one = _plot_bar_or_line("bar", labels, plot_df, series, x, title=title, ylabel=y_label)
            u, n = _save_chart_png(one, f"{stem}-{kind}-{ji}")
            urls.append(u)
            names_out.append(n)
        plot_cols = [str(x), *[str(c) for c in series]]
        plot_rows = list(
            zip(plot_df["label"].tolist(), *[plot_df[c].tolist() for c in series])
        )
        md_chunks.append(
            f"### 图{ji + 1} {kind} {'、'.join(series)}\n"
            + rows_to_markdown(plot_cols, plot_rows, cap=200)
        )
    if not urls:
        raise ValueError("没有生成任何图表")
    n = 0
    try:
        n = int(con.execute(f'SELECT COUNT(*) FROM "{tbl}"').fetchone()[0])
    except Exception:  # noqa: BLE001
        n = 0
    series_note = "、".join(
        str(c)
        for c in (charts[0].get("series") or spec.get("series") or nums)
        if c
    )
    combo_bits = [
        str(j.get("line_mode") or "")
        for j in charts
        if str(j.get("kind") or spec.get("kind") or "") == "combo"
    ]
    kinds_note = ",".join(str(j.get("kind") or spec.get("kind") or "") for j in charts)
    note = (
        f"chart=spec 图数={len(urls)} x={x} 源表行数={n} series={series_note}"
        f" kinds={kinds_note} combo_line={','.join(combo_bits) or '-'}。"
        "图中系列以「作图数据」列名为准：有合计列则与各类商品并列，不是只画总体。"
        "combo_line=total 表示折线是各月合计，不是每类商品一条折线。"
    )
    if "pie" in kinds_note:
        note += " 含饼图 chart=pie。"
    return urls[0], names_out[0], note + "\n\n## 作图数据（全部）\n" + "\n\n".join(md_chunks), urls


def render_chart(
    con,
    table_map: dict[str, str],
    text: str,
    *,
    stem: str = "chart",
) -> tuple[str, str, str, list[str]]:
    """返回 (url, filename, note, all_urls)。"""
    _setup_matplotlib()

    tbl = next(iter(table_map.values()))
    cols = con.execute(f'DESCRIBE "{tbl}"').fetchall()
    names = [c[0] for c in cols]
    types = {c[0]: c[1] for c in cols}
    has_order = "_src_order" in names
    nums = [c for c in names if is_numeric_type(str(types.get(c, ""))) and c != "_src_order"]
    cats = [c for c in names if c not in nums and c != "_src_order"]
    jobs = parse_chart_jobs(text, nums)
    xy_jobs = [j for j in jobs if j["kind"] in {"bar", "line", "combo"}]
    pie_jobs = [j for j in jobs if j["kind"] == "pie"]
    kind = jobs[0]["kind"] if jobs else "bar"
    x, ys = _pick_chart_axes(text, cats, nums)
    y = ys[0] if ys else None
    plot_md = ""
    extra_urls: list[str] = []
    saved_name = ""
    fig = None
    if len(jobs) == 1 and kind == "hist":
        if not y:
            raise ValueError("没有数值列，无法画直方图")
        df = con.execute(f'SELECT "{y}" FROM "{tbl}" WHERE "{y}" IS NOT NULL LIMIT 5000').fetchdf()
        fig, ax = _new_axes((8.2, 4.4))
        ax.hist(df[y].dropna(), bins=20, color="#1F4E79")
        ax.set_xlabel(y)
        ax.set_title(f"{y} 分布")
        plot_md = f"直方图点数={len(df)}"
    elif kind == "scatter":
        if len(nums) < 2:
            raise ValueError("散点图需要至少两列数值")
        y2 = nums[1] if nums[0] == y else nums[0]
        df = con.execute(
            f'SELECT "{y}" AS x, "{y2}" AS y FROM "{tbl}" WHERE "{y}" IS NOT NULL LIMIT 800'
        ).fetchdf()
        fig, ax = _new_axes((8.2, 4.4))
        ax.scatter(df["x"], df["y"], alpha=0.7, c="#1F4E79")
        ax.set_xlabel(y)
        ax.set_ylabel(y2)
        ax.set_title(f"{y} vs {y2}")
        plot_md = f"散点点数={len(df)}"
    elif pie_jobs and not xy_jobs:
        series = ys or nums
        if not series:
            raise ValueError("饼图需要数值列")
        multi = wants_multi_pie(text) and x and len(series) >= 2
        if len(series) >= 2 and x:
            order_sql = 'ORDER BY "_src_order"' if has_order else (
                f'ORDER BY TRY_CAST(regexp_extract(CAST("{x}" AS VARCHAR), \'(\\d+)\', 1) AS INTEGER) '
                f'NULLS LAST, "{x}"'
                if looks_like_month_col(x)
                else f'ORDER BY "{x}"'
            )
            quoted = ", ".join(f'"{c}"' for c in series)
            df = con.execute(
                f'SELECT "{x}" AS label, {quoted} FROM "{tbl}" '
                f'WHERE "{x}" IS NOT NULL {order_sql}'
            ).fetchdf()
            pie_labels = [str(c) for c in series]
            urls: list[str] = []
            names: list[str] = []
            md_chunks: list[str] = []
            if multi:
                for rec in df.itertuples(index=False):
                    row_label = str(rec[0])
                    vals = [float(v or 0) for v in rec[1:]]
                    fig = _draw_pie(pie_labels, vals, f"{row_label} 各商品销量占比")
                    u, n = _save_chart_png(fig, f"{stem}-{row_label}")
                    urls.append(u)
                    names.append(n)
                    md_chunks.append(
                        rows_to_markdown(
                            ["品类", "销量"],
                            list(zip(pie_labels, vals)),
                        )
                    )
                    md_chunks[-1] = f"### {row_label}\n{md_chunks[-1]}"
            totals = [float(df[c].sum()) for c in series]
            fig = _draw_pie(pie_labels, totals, "整体各商品销量占比")
            u, n = _save_chart_png(fig, f"{stem}-整体")
            urls.append(u)
            names.append(n)
            md_chunks.append(
                "### 整体\n"
                + rows_to_markdown(["品类", "销量"], list(zip(pie_labels, totals)))
            )
            extra_urls = urls
            saved_name = names[0]
            plot_md = "\n\n".join(md_chunks)
            fig = None  # already saved
        else:
            if not x:
                raise ValueError("饼图需要类别列")
            metric = f'SUM("{y}")' if y else "COUNT(*)"
            df = con.execute(
                f'SELECT "{x}" AS label, {metric} AS val FROM "{tbl}" '
                f'GROUP BY "{x}" ORDER BY val DESC LIMIT 12'
            ).fetchdf()
            fig, ax = _new_axes((7.2, 6.4))
            ax.pie(df["val"], labels=df["label"].astype(str), autopct="%1.1f%%")
            ax.set_title(str(x))
            plot_md = rows_to_markdown(
                ["类别", "数值"],
                [tuple(r) for r in df.itertuples(index=False, name=None)],
            )
    else:
        if not x:
            raise ValueError("柱状/折线图需要类别列或日期列（如月份）")
        if not nums:
            raise ValueError("没有可绘制的数值列")
        order_sql = 'ORDER BY "_src_order"' if has_order else (
            f'ORDER BY TRY_CAST(regexp_extract(CAST("{x}" AS VARCHAR), \'(\\d+)\', 1) AS INTEGER) '
            f'NULLS LAST, "{x}"'
            if looks_like_month_col(x)
            else f'ORDER BY "{x}"'
        )
        work = xy_jobs or [
            {"kind": kind if kind in {"bar", "line"} else "bar", "ctx": strip_table_from_prompt(text), "split": wants_split_series(text)}
        ]
        quoted = ", ".join(f'"{c}"' for c in nums)
        df = con.execute(
            f'SELECT "{x}" AS label, {quoted} FROM "{tbl}" '
            f'WHERE "{x}" IS NOT NULL {order_sql}'
        ).fetchdf()
        labels = df["label"].astype(str).tolist()
        urls: list[str] = []
        names: list[str] = []
        md_chunks: list[str] = []
        user_title = extract_chart_title(text)
        for ji, job in enumerate(work):
            series, use_total = resolve_job_series(job, nums)
            plot_df = df
            ctx = str(job.get("ctx") or "")
            jkind = str(job.get("kind") or "bar")
            job_title = resolve_chart_title(text, None, series=series, kind=jkind)
            y_label = infer_y_label(text, series)
            if wants_total_as_extra_series(ctx) or job.get("include_total_as_series"):
                plot_df = df.copy()
                plot_df["合计"] = plot_df[nums].sum(axis=1)
                series = [c for c in series if c != "合计"] or list(nums)
                if "合计" not in series:
                    series = list(series) + ["合计"]
                use_total = False
                job_title = user_title or resolve_chart_title(
                    text, "销量情况（含总销量）", series=series, kind=jkind
                )
            if use_total:
                plot_df = df.copy()
                plot_df["合计"] = plot_df[nums].sum(axis=1)
                series = ["合计"]
                job_title = user_title or resolve_chart_title(
                    text, "总体销量（各月合计）", series=series, kind=jkind
                )
            if not series:
                continue
            split = bool(job.get("split")) and len(series) > 1 and jkind != "combo"
            if jkind == "combo":
                one = _plot_combo(
                    labels,
                    plot_df,
                    series,
                    x,
                    title=job_title
                    or (
                        "柱状图 + 合计变化折线"
                        if combo_line_mode(text) == "total"
                        else "柱状图 + 各系列变化折线"
                    ),
                    line_mode=combo_line_mode(text),
                    ylabel=y_label,
                )
                u, n = _save_chart_png(one, f"{stem}-combo")
                urls.append(u)
                names.append(n)
            elif split:
                for col in series:
                    one = _plot_bar_or_line(jkind, labels, plot_df, [col], x, ylabel=y_label)
                    u, n = _save_chart_png(one, f"{stem}-{col}")
                    urls.append(u)
                    names.append(n)
            else:
                one = _plot_bar_or_line(
                    jkind, labels, plot_df, series, x, title=job_title, ylabel=y_label
                )
                u, n = _save_chart_png(one, f"{stem}-{jkind}-{ji}")
                urls.append(u)
                names.append(n)
            plot_cols = [str(x), *[str(c) for c in series]]
            plot_rows = list(
                zip(plot_df["label"].tolist(), *[plot_df[c].tolist() for c in series])
            )
            md_chunks.append(
                f"### 图{ji + 1} {jkind} {'、'.join(series)}\n"
                + rows_to_markdown(plot_cols, plot_rows, cap=200)
            )
        extra_urls = urls
        saved_name = names[0] if names else ""
        fig = None
        if pie_jobs:
            for j in pie_jobs:
                ctx = str(j.get("ctx") or "")
                named = [c for c in nums if c and str(c) in ctx]
                series_p = (
                    list(nums)
                    if _is_total_ctx(ctx) or "占比" in ctx or _is_monthly_facet(ctx) or not named
                    else named
                )
                monthly = _is_monthly_facet(ctx)
                facet = monthly or (wants_multi_pie(ctx) and not _is_total_ctx(ctx))
                include_overall = (not monthly) or ("整体" in ctx)
                pu, pn, pmd = _pies_from_wide_df(
                    df[["label", *series_p]] if "label" in df.columns else df,
                    series_p,
                    stem,
                    facet=facet,
                    include_overall=include_overall,
                    overall_title=infer_chart_title(text, series=series_p, kind="pie") or "占比",
                )
                urls.extend(pu)
                names.extend(pn)
                md_chunks.extend(pmd)
            extra_urls = urls
            saved_name = names[0] if names else saved_name
        plot_md = "\n\n".join(md_chunks)
    if fig is not None:
        url, name = _save_chart_png(fig, stem)
        extra_urls = [url]
        saved_name = name
    elif extra_urls:
        url, name = extra_urls[0], saved_name
    else:
        raise ValueError("没有生成任何图表")
    n = 0
    try:
        n = int(con.execute(f'SELECT COUNT(*) FROM "{tbl}"').fetchone()[0])
    except Exception:  # noqa: BLE001
        n = 0
    n_img = len(extra_urls)
    if kind == "pie" and n_img > 1:
        note = (
            f"chart=pie 图数={n_img} series={','.join(ys or [])} 源表行数={n}。"
            "已按每个类别（如月份）各画一张商品占比饼图，并另画一张整体占比。"
            "饼图没有横轴；禁止把月份当成唯一切片。"
        )
    elif n_img > 1:
        note = (
            f"jobs={len(jobs)} 图数={n_img} x={x} 源表行数={n}。"
            "已按用户一句话里的多种图（如折线 + 总体柱状）分别出图。"
            "禁止说脚本只能生成一种图；禁止让用户重跑或填写 chart=/series= 参数。"
        )
    else:
        note = (
            f"chart={kind} x={x} series={','.join(ys) if ys else y} 源表行数={n}。"
            f"未指定品类时绘制全部数值列；横轴是「{x}」。"
        )
    return url, name, note + "\n\n## 作图数据（全部）\n" + plot_md, extra_urls
