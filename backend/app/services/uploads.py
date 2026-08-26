"""上传文件落盘（文本类小文件）。"""
from __future__ import annotations

import re
import uuid
from pathlib import Path

from fastapi import UploadFile

ROOT = Path(__file__).resolve().parents[4]
UPLOAD_ROOT = ROOT / "uploads"
ALLOWED_SUFFIX = {
    ".csv",
    ".json",
    ".txt",
    ".tsv",
    ".md",
    ".xlsx",
    ".xls",
    ".pdf",
    ".doc",
    ".docx",
    ".ppt",
    ".pptx",
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".html",
    ".htm",
    ".rtf",
    ".odt",
}
MAX_FILE_BYTES = 32 * 1024 * 1024
MAX_FILES = 9
PREVIEW_CHARS = 4000
CSV_PREVIEW_CHARS = 24000
CURRENT_FILE_MARKER = "当前上传文件（本次任务的唯一数据源）"


def _safe_name(name: str) -> str:
    base = Path(name or "file.txt").name
    base = re.sub(r"[^\w.\-一-龥]+", "_", base)
    return base[:120] or "file.txt"


def _preview_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return "（未安装 pypdf，无法预览 PDF 文本；请 pip install pypdf）"
    try:
        reader = PdfReader(str(path))
        parts: list[str] = [f"## PDF pages: {len(reader.pages)}"]
        for i, page in enumerate(reader.pages[:8]):
            text = (page.extract_text() or "").strip()
            if not text:
                parts.append(f"### Page {i + 1}\n（本页无抽取到文本，可能是扫描件）")
                continue
            parts.append(f"### Page {i + 1}\n{text}")
            if sum(len(p) for p in parts) >= PREVIEW_CHARS:
                parts.append("…(截断)")
                break
        return "\n\n".join(parts)[:PREVIEW_CHARS]
    except Exception as exc:  # noqa: BLE001
        return f"（无法解析 PDF 预览: {exc}）"


def _preview_bytes(path: Path, data: bytes) -> str:
    suffix = path.suffix.lower()
    if suffix in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}:
        return f"（图像 {path.name}，{len(data)} bytes）"
    if suffix in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            parts: list[str] = []
            for sheet in wb.worksheets[:4]:
                parts.append(f"## Sheet: {sheet.title}")
                rows_out = 0
                for row in sheet.iter_rows(values_only=True):
                    cells = ["" if c is None else str(c) for c in row]
                    if not any(x.strip() for x in cells):
                        continue
                    parts.append("\t".join(cells))
                    rows_out += 1
                    if rows_out >= 40:
                        parts.append("…(截断)")
                        break
            wb.close()
            text = "\n".join(parts)
            return text[:PREVIEW_CHARS]
        except Exception as exc:  # noqa: BLE001
            return f"（无法解析 Excel 预览: {exc}）"
    if suffix == ".pdf":
        return _preview_pdf(path)
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("utf-8", errors="replace")
    cap = CSV_PREVIEW_CHARS if suffix in {".csv", ".tsv"} else PREVIEW_CHARS
    return text[:cap]


def collect_upload_files(form) -> list:
    """从 multipart form 收集上传文件（兼容 files/file；不依赖严格 isinstance）。"""
    out = []
    for key, value in form.multi_items():
        if key not in {"files", "file"}:
            continue
        filename = getattr(value, "filename", None)
        read = getattr(value, "read", None)
        if filename and callable(read):
            out.append(value)
    return out


async def save_upload_files(
    user_id: str,
    files: list[UploadFile] | None,
) -> list[dict]:
    """保存上传文件，返回 [{name,path,size,preview}]。"""
    if not files:
        return []
    if len(files) > MAX_FILES:
        raise ValueError(f"最多上传 {MAX_FILES} 个文件")
    call_dir = UPLOAD_ROOT / (user_id or "anon") / uuid.uuid4().hex
    call_dir.mkdir(parents=True, exist_ok=True)
    saved: list[dict] = []
    for f in files:
        if not f or not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower()
        if suffix not in ALLOWED_SUFFIX:
            raise ValueError(f"不支持的文件类型: {suffix or '(无后缀)'}")
        data = await f.read()
        if len(data) > MAX_FILE_BYTES:
            raise ValueError(f"文件过大（>{MAX_FILE_BYTES} bytes）: {f.filename}")
        path = call_dir / _safe_name(f.filename)
        path.write_bytes(data)
        preview = _preview_bytes(path, data)
        saved.append(
            {
                "name": path.name,
                "path": str(path.resolve()),
                "size": len(data),
                "preview": preview,
            }
        )
    return saved


def format_current_file_context(files: list[dict] | None) -> str:
    """给模型的「本次上传」块：有文件才返回，强调不得用历史里的旧文件。"""
    if not files:
        return ""
    names = [str(f.get("name") or "").strip() for f in files if f.get("name")]
    names = [n for n in names if n]
    name_list = "、".join(names) if names else "（未命名）"
    return (
        f"## {CURRENT_FILE_MARKER}\n"
        f"本次用户上传的文件：{name_list}。必须只分析这些文件的内容。"
        "禁止把先前对话里出现的其它文件名或旧报告当作本次数据来源，"
        "禁止编造与下列内容不符的指标、样本或文件名。\n\n"
        f"{format_file_summary(files)}"
    )


def inject_current_files(user_text: str, files: list[dict] | None) -> str:
    """自定义 Skill 模板常只有 {{input}}，把上传文件强制写入本次 user 消息。"""
    if not files:
        return user_text or ""
    text = (user_text or "").rstrip()
    followup = any(
        k in text for k in ("上文", "上面", "上述", "## 先前对话")
    )
    if followup:
        summary = format_file_summary(files)
        if summary and summary not in text:
            return f"{text}\n\n{summary}"
        return text
    if CURRENT_FILE_MARKER in text:
        return text
    summary = format_file_summary(files)
    if summary and summary in text:
        return (
            f"{text}\n\n必须只分析上述上传文件；忽略历史对话中的其它文件名与旧报告。"
        )
    ctx = format_current_file_context(files)
    return f"{text}\n\n{ctx}".strip() if text else ctx


def format_file_summary(files: list[dict] | None) -> str:
    if not files:
        return "（无上传文件）"
    parts = [
        "以下文件已由平台接收并保存在服务器，可直接用于处理（勿再向用户索要本地路径）："
    ]
    for f in files:
        name = f.get("name") or ""
        size = f.get("size")
        path = f.get("path") or ""
        preview = f.get("preview") or ""
        parts.append(
            f"- name: {name}\n"
            f"  size: {size} bytes\n"
            f"  serverPath: {path}\n"
            f"  preview:\n```\n{preview}\n```"
        )
    return "\n".join(parts)
